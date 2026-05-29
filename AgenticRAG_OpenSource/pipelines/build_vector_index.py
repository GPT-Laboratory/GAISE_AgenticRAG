"""Build the local Chroma vector index from knowledge-base files in data/raw/.

Pipeline:  data/raw/* -> parse per filetype -> chunk (~400 tokens, ~60 overlap)
           -> e5 "passage: " embeddings (normalized) -> Chroma (cosine, persisted).

Idempotent: the collection is deleted and rebuilt on every run, so re-running
never duplicates chunks. Numeric processed CSVs are NOT indexed (they live behind
the analytics tools).

Usage:  python pipelines/build_vector_index.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Iterable, Optional

# Allow running as a plain script (python pipelines/build_vector_index.py)
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import settings  # noqa: E402

RAW_DIR = Path(settings.VECTOR_STORE_DIR).resolve().parent.parent / "raw"
# VECTOR_STORE_DIR is .../data/cache/vector_index ; raw is .../data/raw
RAW_DIR = (Path(settings.VECTOR_STORE_DIR).resolve().parents[1] / "raw")

# e5 prefix conventions — keep in sync with analytics/document_search.py
PASSAGE_PREFIX = "passage: "

# Approx token budget per chunk (e5 char heuristic: ~4 chars/token)
CHUNK_CHARS = 1600          # ~400 tokens
CHUNK_OVERLAP_CHARS = 240   # ~60 tokens
MIN_CHUNK_CHARS = 200       # drop fragments shorter than ~50 tokens

INDEXABLE_SUFFIXES = {".pdf", ".txt", ".md", ".docx", ".xlsx"}

# Files kept in data/raw/ for reference but deliberately NOT indexed.
# sample_FAQs.txt is the project's example-question list (a spec, not knowledge);
# the definitions it references already live in methodology.md, so indexing it
# would only add retrieval noise.
SKIP_FILES = {"sample_FAQs.txt"}


# --------------------------------------------------------------------------- #
# Parsers — each yields (text, page, sheet) records
# --------------------------------------------------------------------------- #
def _parse_pdf(path: Path) -> Iterable[tuple[str, Optional[int], Optional[str]]]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    for i, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            yield text, i, None


def _parse_text(path: Path) -> Iterable[tuple[str, Optional[int], Optional[str]]]:
    text = path.read_text(encoding="utf-8", errors="ignore").strip()
    if text:
        yield text, None, None


def _parse_docx(path: Path) -> Iterable[tuple[str, Optional[int], Optional[str]]]:
    from docx import Document

    doc = Document(str(path))
    parts = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
    text = "\n".join(parts).strip()
    if text:
        yield text, None, None


def _parse_xlsx(path: Path) -> Iterable[tuple[str, Optional[int], Optional[str]]]:
    """Extract only non-numeric text (titles, headers, notes). Numeric tables
    belong to the analytics CSVs and are intentionally skipped."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    for sheet in wb.worksheets:
        texts: list[str] = [sheet.title]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    s = cell.strip()
                    # keep textual labels/notes; skip pure numbers / stringified numbers
                    if s and not _looks_numeric(s):
                        texts.append(s)
        joined = "\n".join(dict.fromkeys(texts)).strip()  # de-dup, preserve order
        if len(joined) > len(sheet.title):  # more than just the title
            yield joined, None, sheet.title
    wb.close()


def _looks_numeric(s: str) -> bool:
    try:
        float(s.replace(",", "").replace(" ", ""))
        return True
    except ValueError:
        return False


PARSERS = {
    ".pdf": _parse_pdf,
    ".txt": _parse_text,
    ".md": _parse_text,
    ".docx": _parse_docx,
    ".xlsx": _parse_xlsx,
}


# --------------------------------------------------------------------------- #
# Chunking
# --------------------------------------------------------------------------- #
def _chunk(text: str) -> list[str]:
    text = " ".join(text.split())  # normalize whitespace
    if len(text) <= CHUNK_CHARS:
        return [text] if len(text) >= MIN_CHUNK_CHARS else ([text] if text else [])

    chunks: list[str] = []
    start = 0
    n = len(text)
    while start < n:
        end = min(start + CHUNK_CHARS, n)
        # try to break on a sentence/space boundary near the end
        if end < n:
            window = text.rfind(" ", start + MIN_CHUNK_CHARS, end)
            if window != -1:
                end = window
        chunk = text[start:end].strip()
        if len(chunk) >= MIN_CHUNK_CHARS or (chunk and not chunks):
            chunks.append(chunk)
        if end >= n:
            break
        start = max(end - CHUNK_OVERLAP_CHARS, start + 1)
    return chunks


# --------------------------------------------------------------------------- #
# Build
# --------------------------------------------------------------------------- #
def build() -> None:
    import chromadb
    from sentence_transformers import SentenceTransformer

    if not RAW_DIR.exists():
        raise SystemExit(f"Raw KB directory not found: {RAW_DIR}")

    files = sorted(
        p
        for p in RAW_DIR.iterdir()
        if p.suffix.lower() in INDEXABLE_SUFFIXES and p.name not in SKIP_FILES
    )
    if not files:
        raise SystemExit(f"No indexable files in {RAW_DIR} (looked for {sorted(INDEXABLE_SUFFIXES)})")

    print(f"Loading embedding model: {settings.EMBEDDING_MODEL} ...")
    model = SentenceTransformer(settings.EMBEDDING_MODEL)

    ids: list[str] = []
    docs: list[str] = []
    metas: list[dict[str, Any]] = []
    per_source: dict[str, int] = {}

    for path in files:
        parser = PARSERS[path.suffix.lower()]
        chunk_index = 0
        try:
            records = list(parser(path))
        except Exception as exc:
            print(f"  ! skipping {path.name}: {exc}")
            continue
        for text, page, sheet in records:
            for chunk in _chunk(text):
                ids.append(f"{path.name}::{chunk_index}")
                docs.append(chunk)
                metas.append(
                    {
                        "source": path.name,
                        "page": page if page is not None else "",
                        "sheet": sheet if sheet is not None else "",
                        "chunk_index": chunk_index,
                    }
                )
                chunk_index += 1
        per_source[path.name] = chunk_index
        print(f"  {path.name}: {chunk_index} chunks")

    if not docs:
        raise SystemExit("No chunks produced — nothing to index.")

    print(f"Embedding {len(docs)} chunks ...")
    embeddings = model.encode(
        [PASSAGE_PREFIX + d for d in docs],
        normalize_embeddings=True,
        show_progress_bar=True,
        batch_size=32,
    )

    print(f"Persisting to Chroma at {settings.VECTOR_STORE_DIR} ...")
    client = chromadb.PersistentClient(path=settings.VECTOR_STORE_DIR)
    try:
        client.delete_collection(settings.VECTOR_COLLECTION)  # idempotent rebuild
    except Exception:
        pass
    collection = client.create_collection(
        settings.VECTOR_COLLECTION, metadata={"hnsw:space": "cosine"}
    )
    collection.add(
        ids=ids,
        embeddings=[e.tolist() for e in embeddings],
        documents=docs,
        metadatas=metas,
    )

    total = sum(per_source.values())
    print("\nDone.")
    print(f"  collection: {settings.VECTOR_COLLECTION}")
    print(f"  total chunks: {total}")
    for src, n in per_source.items():
        print(f"    - {src}: {n}")


if __name__ == "__main__":
    build()
